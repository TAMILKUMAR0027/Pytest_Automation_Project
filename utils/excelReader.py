from openpyxl import Workbook
import openpyxl
import os


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

SEARCH_EXCEL_PATH = os.path.join(
    BASE_DIR, "data_provider","searchProduct.xlsx"   
)

SEARCH_SHEET = "SearchData"


def get_data(path, sheet_name):
    final_list = []
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheet_name]
    total_row = sheet.max_row
    total_columns = sheet.max_column

    for r in range(2, total_row + 1):
        row_list = []
        for c in range(1, total_columns + 1):
            row_list.append(sheet.cell(r, c).value)
        final_list.append(row_list)
    return final_list


def get_registration_data(path, sheet_name):
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheet_name]

    return {
        "firstname":        sheet.cell(2, 1).value,
        "lastname":         sheet.cell(2, 2).value,
        "email":            sheet.cell(2, 3).value,
        "telephone":        sheet.cell(2, 4).value,
        "password":         sheet.cell(2, 5).value,
        "confirm_password": sheet.cell(2, 6).value,
        "company":          sheet.cell(2, 7).value,
        "address1":         sheet.cell(2, 8).value,
        "address2":         sheet.cell(2, 9).value,
        "city":             sheet.cell(2, 10).value,
        "postcode":         sheet.cell(2, 11).value,
        "country":          sheet.cell(2, 12).value,
        "region":           sheet.cell(2, 13).value,
    }